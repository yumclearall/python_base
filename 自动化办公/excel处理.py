from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

wb = load_workbook('./高效办公/考勤数据表.xlsx')
ws = wb.active

late_font = Font(color='FF8C00') # 深橙色
leave_font = Font(color='FF0000') # 纯红色
absenteeism_font = Font(bold=True) # 加粗

# 定义旷工的样式
absenteeism_fill = PatternFill(patternType='solid', fgColor='FF6347')


for row in ws.iter_rows(min_row=3, values_only=False):
    for cell in row:
        for row in ws.iter_rows(min_row=3, values_only=False):attendance_count = 0
    late_count = 0
    leave_count = 0
    absenteeism_count = 0

    for cell in row:
        if cell.value == '出勤':
            attendance_count += 1
        elif cell.value == '迟到':
            late_count += 1
            cell.font = late_font
        elif cell.value == '请假':
            leave_count += 1
            cell.font = leave_font
        elif cell.value == '旷工':
            absenteeism_count += 1
            cell.fill = absenteeism_fill


n = 3
for row in ws.iter_rows(min_row=3, values_only=False):
    attendance_count = 0
    late_count = 0
    leave_count = 0
    absenteeism_count = 0

    for cell in row:
        if cell.value == '出勤':
            attendance_count += 1
        elif cell.value == '迟到':
            late_count += 1
            cell.font = late_font
        elif cell.value == '请假':
            leave_count += 1
            cell.font = leave_font
        elif cell.value == '旷工':
            absenteeism_count += 1
            cell.fill = absenteeism_fill

    ws['W' + str(n)] = attendance_count
    ws['X' + str(n)] = late_count
    ws['Y' + str(n)] = leave_count
    ws['Z' + str(n)] = absenteeism_count
    n += 1

    
ws['X1'] = '助教'

wb.save('./高效办公/考勤数据表.xlsx')