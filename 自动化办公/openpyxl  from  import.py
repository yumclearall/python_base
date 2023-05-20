import openpyxl

# 通过文件路径，打开工作簿
wb1 = openpyxl.load_workbook('c:/Users/86156/Desktop/.py/自动化办公/openpyxl/abcd.xlsx')
# 用 Workbook() 创建新工作簿
wb2 = openpyxl.Workbook()
wb2.save('c:/Users/86156/Desktop/.py/自动化办公/openpyxl/abcd.xlsx')
print(wb2)

from openpyxl import load_workbook, Workbook

# 通过文件路径，打开已有工作簿
wb1 = load_workbook('c:/Users/86156/Desktop/.py/自动化办公/openpyxl/abcd.xlsx')
# 用 Workbook() 创建新工作簿
wb2 = Workbook()
wb2.save('c:/Users/86156/Desktop/.py/自动化办公/openpyxl/abcd.xlsx')
print(wb2)