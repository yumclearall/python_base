# 课后练习：取数汇总并写入
from openpyxl import load_workbook
wz='openpyxl/公司人员名单.xlsx'
bc='openpyxl/江宇工资信息表.xlsx'
wzs=load_workbook(wz)
wzzs=wzs.active
for row in wzzs.iter_rows(min_row=2,values_only=True):
    if row[1]=='江宇':
        row[6]=row[3]+row[4]+row[5]
        wzs2=load_workbook(bc)
        bcd=bc.active
        bcd['H1']='总工资'
        for i in range(3,6):
            bcd.append(row[i])
        wzs2.save