# 从openpyxl库导入load_workbook和Workbook
from openpyxl import load_workbook, Workbook

# 打开【10月员工绩效表.xlsx】工作簿
performance_wb = load_workbook('openpyxl/10月员工绩效表.xlsx')
# 获取活动工作表
performance_ws = performance_wb.active

# 新建工作簿
new_wb = Workbook()
# 获取活动工作表
new_ws = new_wb.active

# 获取performance_ws的前十行数据
for row in performance_ws.iter_rows(max_row=10, values_only=True):
    # 将数据写入新的工作表
    new_ws.append(row)

# 保存新工作簿为【员工绩效表-模板.xlsx】
new_wb.save('openpyxl/10月员工绩效表-模板.xlsx')