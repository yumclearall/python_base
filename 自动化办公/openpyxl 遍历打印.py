# 从openpyxl库导入load_workbook函数
from openpyxl import load_workbook

# 打开【10月员工绩效表.xlsx】工作簿
performance_wb = load_workbook('openpyxl/10月员工绩效表.xlsx')
# 获取活动工作表
performance_ws = performance_wb.active

# 创建员工信息字典
staff_info = {}

# 从第二行开始读取工作表中的信息
for row in performance_ws.iter_rows(min_row=2, values_only=True):
    # 取出工号
    member_number = row[0]
    # 将信息存入员工信息字典
    staff_info[member_number] = {
         '姓名': row[1],
         '部门': row[2],
         '绩效': row[3],
         '奖金': row[4],
         '基本工资': row[5],
         '是否确认': row[6]
     }
print(staff_info)