# 从openpyxl库导入load_workbook和Workbook
from openpyxl import load_workbook, Workbook

# 打开【10月员工绩效表.xlsx】工作簿
performance_wb = load_workbook('openpyxl/10月员工绩效表.xlsx')
# 获取活动工作表
performance_ws = performance_wb.active

# 获取performance_ws中除表头外的数据
for row in performance_ws.iter_rows(min_row=2, values_only=True):
    # 读取【工号】
    staff_id = row[0]
    # 读取【员工姓名】
    staff_name = row[1]
    # 读取【绩效】
    performance = row[3]
    # 读取【提成】
    bonus = row[4]
    # 计算“奖金”
    award = performance + bonus
    # 打印结果
    print('工号：{}，姓名：{}，本月奖金为：{}'.format(staff_id, staff_name, award))

