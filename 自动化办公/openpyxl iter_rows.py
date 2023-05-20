from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('c:/Users/86156/Desktop/.py/自动化办公/openpyxl/公司人员名单.xlsx')
# 获取【'上半年公司名单'】工作表
fhy_ws = staff_wb['上半年公司名单']

# 返回第2行至第12行，第2列（B列）至第3列（C列）这个范围的单元格内的所有数据（值）
for row in fhy_ws.iter_rows(min_row=2, max_row=10, min_col=2, max_col=3, values_only=True):
    print(row)
for row in fhy_ws.iter_rows(min_row=2, max_row=10, min_col=2, max_col=3):
    print(row)