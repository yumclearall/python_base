# 导入模块
from openpyxl import load_workbook

# 获取数据
wb  = load_workbook('openpyxl/10月员工绩效表.xlsx')
ws = wb.active

performance = ws['E11'].value
bonus = ws['F11'].value
base = ws['D11'].value

# 使用数据
salary = performance + bonus + base
wb2=load_workbook('openpyxl/江宇工资信息表.xlsx')
wb2s=wb2.active
# 输出结果
wb2s['H1'].value = '总工资'
wb2s['H11'].value = salary
wb2.save('openpyxl/江宇工资信息表.xlsx')