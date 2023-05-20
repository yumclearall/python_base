# 导入模块
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Side, Border

# 定义表头颜色样式为橙色
tou=PatternFill('solid',fgColor='FF7F24')
# 定义表中颜色样式为淡黄色
zhong=PatternFill('solid',fgColor='FFFFE0')
# 定义表尾颜色样式为淡桔红色
wei=PatternFill('solid',fgColor='EE9572')

# 定义对齐样式横向居中、纵向居中
duiqi=Alignment(horizontal='center',vertical='center')

# 定义边样式为细条
side=Side('thin')
# 定义表头边框样式，有底边和右边
touk=Border(bottom=side,right=side)
# 定义表中、表尾边框样式，有左边
zhongk=Border(left=side)

# 设置文件夹路径
lujing='excel样式/各部门利润表汇总-副件/'
# 返回当前目录下所有文件名
lujing2=os.listdir(lujing)

# 循环文件名列表
for lj in lujing2:
    # 拼接文件路径
    lujing3=lujing+lj
    # 打开工作簿
    wb=load_workbook(lujing3)
    # 打开工作表
    ws=wb.active

    # 调整列宽
    ws.column_dimensions['A'].width=15
    ws.column_dimensions['B'].width=20
    ws.column_dimensions['C'].width=25
    ws.column_dimensions['D'].width=20
    ws.column_dimensions['E'].width=15
    ws.column_dimensions['F'].width=15

    # 循环第一行单元格，调整表头样式
    for row in ws[1]:
        # 设置单元格填充颜色
        row.fill=tou
        # 设置单元格对齐方式
        row.alignment=duiqi
        # 设置单元格边框
        row.border=touk

    # 获取最后一行行号
    weis=ws.max_row

    # 从第二行开始，循环到倒数第二行
    for row in ws.iter_rows(min_row=2,max_row=weis-1):
        # 循环取出单元格，调整表中样式
        for cell in row:
            cell.fill=zhong
            cell.alignment=duiqi
            cell.border=zhongk



    # 循环最后一行单元格，调整表尾样式
    for row in ws[weis]:
        row.fill=wei
        row.alignment=duiqi
        row.border=zhongk



    # 保存
    wb.save(lujing3)