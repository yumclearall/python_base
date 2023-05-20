# 从openpyxl库styles模块中导入四个类
from openpyxl.styles import PatternFill, Alignment, Side, Border

# 定义表头填充样式，颜色为橙色，纯色填充
header_fill = PatternFill('solid', fgColor='FF7F24')
# 定义表中填充样式，颜色为淡黄色，纯色填充
content_fill = PatternFill('solid', fgColor='FFFFE0')
# 定义表尾填充样式，颜色为淡桔红色，纯色填充
bottom_fill = PatternFill('solid', fgColor='EE9572')

# 定义对齐样式横向居中、纵向居中
align = Alignment(horizontal='center', vertical='center')

# 定义边样式为细边框
side = Side('thin')
# 定义表头边框样式，有底部和右侧细边框
header_border = Border(bottom=side, right=side)
# 定义表中、表尾边框样式，有左侧细边框
content_border = Border(left=side)