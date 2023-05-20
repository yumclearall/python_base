from openpyxl import load_workbook

# 打开【pre_practice2.xlsx】工作簿
wb = load_workbook('openpyxl/10月员工绩效表-模板.xlsx')
# 获取活动工作表
ws = wb.active

sum = 0
# 获取基本工资（D）列的数据（不包含表头），并计算总和
for row in ws['F']:
    if row.value=='基本工资':
        continue
    sum=sum+row.value
        

# 打印输出结果
print('表格中员工的基本工资总和是：{}'.format(sum))


from openpyxl import load_workbook

# 打开【pre_practice2.xlsx】工作簿
wb = load_workbook('openpyxl/10月员工绩效表-模板.xlsx')
# 获取活动工作表
ws = wb.active

sum = 0

# 获取基本工资（D）列的数据（不包含表头），并计算总和
for row in ws.iter_rows(min_row=2, values_only=True):
    sum += row[5]

# 打印输出结果
print('表格中员工的基本工资总和是：{}'.format(sum))